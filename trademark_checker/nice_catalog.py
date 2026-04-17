from __future__ import annotations

import json
import re
import zipfile
from functools import lru_cache
from pathlib import Path
from typing import Callable, Iterable
from xml.etree import ElementTree as ET

from similarity_code_db import (
    ALIAS_TABLE_PATH,
    LEARNED_MATCHES_PATH,
    SIMILARITY_CODE_SOURCE_PATH,
    derive_similarity_mapping,
)

ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = Path(__file__).resolve().parent / "data"
DOCS_DIR = ROOT_DIR / "docs"
EXCEL_SOURCE_PATH = DOCS_DIR / "지식재산처_상품분류_니스분류.xlsx"
CLASS_CATALOG_PATH = DATA_DIR / "nice_class_catalog.json"
GROUP_CATALOG_PATH = DATA_DIR / "nice_group_catalog.json"

XML_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


GOODS_CATEGORY_DEFS = [
    {"group_id": "fashion_accessories", "group_label": "패션의류/잡화", "group_hint": "의류/주얼리"},
    {"group_id": "beauty", "group_label": "뷰티", "group_hint": "화장품/향수"},
    {"group_id": "food", "group_label": "식품", "group_hint": "식품/음료"},
    {"group_id": "sports_auto", "group_label": "스포츠/레저/자동차", "group_hint": "레저/차량"},
    {"group_id": "baby_kids", "group_label": "출산/유아동", "group_hint": "유아/장난감"},
    {"group_id": "home_interior", "group_label": "가구/인테리어", "group_hint": "가구/홈데코"},
    {"group_id": "living_health", "group_label": "생활/건강", "group_hint": "청소/위생/건강"},
    {"group_id": "digital_devices", "group_label": "가전/디지털", "group_hint": "전자기기/가전"},
    {"group_id": "books_hobby_pet", "group_label": "도서/취미/펫", "group_hint": "문구/취미/반려"},
    {"group_id": "industrial_goods", "group_label": "산업", "group_hint": "원료/산업재"},
    {"group_id": "software", "group_label": "소프트웨어", "group_hint": "앱/디지털콘텐츠"},
]

SERVICES_CATEGORY_DEFS = [
    {"group_id": "food_hospitality", "group_label": "요식업/식품", "group_hint": "식당/카페"},
    {"group_id": "fashion_shopping", "group_label": "패션/쇼핑몰", "group_hint": "소매/도매"},
    {"group_id": "beauty_care", "group_label": "뷰티/미용", "group_hint": "미용/관리"},
    {"group_id": "medical_welfare", "group_label": "의료/제약/복지", "group_hint": "의료/복지"},
    {"group_id": "culture_travel_sports", "group_label": "문화/여행/체육", "group_hint": "여행/문화"},
    {"group_id": "daily_convenience", "group_label": "생활/편의서비스", "group_hint": "수리/제작"},
    {"group_id": "education_kids_pet", "group_label": "교육/유아/반려동물", "group_hint": "교육/유아/펫"},
    {"group_id": "home_living_services", "group_label": "홈/리빙", "group_hint": "건설/인테리어"},
    {"group_id": "creator_freelance", "group_label": "크리에이터/프리랜서", "group_hint": "광고/콘텐츠"},
    {"group_id": "it_platform_app", "group_label": "IT/플랫폼/APP", "group_hint": "SaaS/플랫폼"},
    {"group_id": "misc_services", "group_label": "기타 서비스", "group_hint": "금융/가상화폐/법률"},
]

CATEGORY_DEFS = {"goods": GOODS_CATEGORY_DEFS, "services": SERVICES_CATEGORY_DEFS}
CATEGORY_META = {
    kind: {row["group_id"]: row for row in rows}
    for kind, rows in CATEGORY_DEFS.items()
}
SCOPE_SUBSTEP_GROUP = "group"
SCOPE_SUBSTEP_SUBGROUP = "subgroup"
SCOPE_SUBSTEP_REVIEW_READY = "review_ready"


def _load_json(path: Path):
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def _write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def dedupe_strings(values: Iterable[str]) -> list[str]:
    seen = set()
    items: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        items.append(text)
    return items


def dedupe_ints(values: Iterable[int | str]) -> list[int]:
    seen = set()
    items: list[int] = []
    for value in values:
        try:
            number = int(value)
        except (TypeError, ValueError):
            continue
        if number in seen:
            continue
        seen.add(number)
        items.append(number)
    return sorted(items)


def format_nice_class(number: int | str) -> str:
    return f"제{int(number)}류"


def format_nice_classes(numbers: Iterable[int | str]) -> str:
    classes = dedupe_ints(numbers)
    return ", ".join(format_nice_class(number) for number in classes)


def _normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _column_name(cell_ref: str) -> str:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    return match.group(1) if match else ""


def _extract_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall("main:si", XML_NS):
        fragments = [node.text or "" for node in item.findall(".//main:t", XML_NS)]
        values.append("".join(fragments))
    return values


def _read_xml_sheet_rows(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _extract_shared_strings(archive)
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))

    rows: list[dict[str, str]] = []
    for row in root.findall(".//main:sheetData/main:row", XML_NS):
        current: dict[str, str] = {}
        for cell in row.findall("main:c", XML_NS):
            ref = cell.attrib.get("r", "")
            column = _column_name(ref)
            cell_type = cell.attrib.get("t")
            value = ""
            if cell_type == "inlineStr":
                value = "".join(node.text or "" for node in cell.findall(".//main:t", XML_NS))
            else:
                node = cell.find("main:v", XML_NS)
                raw = node.text if node is not None else ""
                if cell_type == "s" and raw:
                    value = shared_strings[int(raw)]
                else:
                    value = raw or ""
            if column:
                current[column] = _normalize_text(value)
        rows.append(current)
    return rows


def _read_openpyxl_rows(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows():
        current: dict[str, str] = {}
        for cell in row:
            value = _normalize_text(cell.value)
            if value:
                current[_column_name(cell.coordinate)] = value
        rows.append(current)
    workbook.close()
    return rows


def _read_excel_rows(path: Path) -> list[dict[str, str]]:
    try:
        return _read_openpyxl_rows(path)
    except Exception:
        return _read_xml_sheet_rows(path)


def _parse_class_number(text: str) -> int | None:
    match = re.search(r"(\d+)", str(text or ""))
    return int(match.group(1)) if match else None


def _split_subgroups(heading: str) -> list[str]:
    cleaned = _normalize_text(heading)
    if not cleaned:
        return []
    parts = re.split(r"\s*;\s*", cleaned)
    return [part.strip(" .") for part in parts if part.strip(" .")]


def _compact_phrase(text: str) -> str:
    value = _normalize_text(text)
    if not value:
        return ""
    value = value.replace("기타 ", "")
    for delimiter in (",", " 및 ", " 또는 ", "와 ", "과 "):
        if delimiter in value:
            head = value.split(delimiter, 1)[0].strip()
            if head:
                return head
    return value


def _build_keywords(group_heading: str, subgroup_label: str) -> list[str]:
    values = [subgroup_label, group_heading, _compact_phrase(subgroup_label)]
    values.extend(re.split(r"[,/·()]| 및 | 또는 |\s", subgroup_label))
    return dedupe_strings(item.strip() for item in values if item and len(item.strip()) >= 2)


def _derive_subgroup_similarity_mapping(kind: str, class_no: int, subgroup_label: str, group_heading: str) -> dict:
    keywords = [
        keyword
        for keyword in _build_keywords(group_heading, subgroup_label)
        if ";" not in keyword and len(keyword) <= 40
    ]
    return derive_similarity_mapping(
        subgroup_label,
        class_no=class_no,
        seed_classes=[class_no],
        seed_keywords=keywords,
    )


def _contains_any(text: str, terms: Iterable[str]) -> bool:
    lowered = text.lower()
    return any(term.lower() in lowered for term in terms)


def _goods_category_id(class_no: int, subgroup_label: str, heading: str) -> str:
    label_text = subgroup_label
    full_text = f"{subgroup_label} {heading}"
    if _contains_any(label_text, ["유아", "아기", "장난감"]):
        return "baby_kids"
    if class_no == 9 and _contains_any(label_text, ["소프트웨어", "앱", "응용프로그램", "애플리케이션", "멀티미디어", "컴퓨터"]):
        return "software"
    if class_no == 3 and _contains_any(label_text, ["화장품", "세면용품", "향수", "치약"]):
        return "beauty"
    if class_no in {14, 18, 25, 26}:
        return "fashion_accessories"
    if class_no in {29, 30, 32, 33, 34}:
        return "food"
    if class_no == 31 and _contains_any(full_text, ["반려", "애완", "사료"]):
        return "books_hobby_pet"
    if class_no == 31:
        return "food"
    if class_no in {12, 13, 22, 28}:
        return "sports_auto"
    if class_no in {19, 20, 24, 27}:
        return "home_interior"
    if class_no in {3, 5, 10, 21}:
        return "living_health"
    if class_no in {9, 11}:
        return "digital_devices"
    if class_no in {15, 16, 23}:
        return "books_hobby_pet"
    return "industrial_goods"


def _services_category_id(class_no: int, subgroup_label: str, heading: str) -> str:
    label_text = subgroup_label
    full_text = f"{subgroup_label} {heading}"
    if class_no == 35:
        if _contains_any(label_text, ["소매", "도매", "판매", "쇼핑", "주문"]):
            return "fashion_shopping"
        if _contains_any(label_text, ["플랫폼", "온라인", "앱", "데이터"]):
            return "it_platform_app"
        return "creator_freelance"
    if class_no == 36:
        return "misc_services"
    if class_no == 37:
        return "home_living_services"
    if class_no == 38:
        return "it_platform_app"
    if class_no == 39:
        if _contains_any(label_text, ["여행", "관광"]):
            return "culture_travel_sports"
        return "daily_convenience"
    if class_no == 40:
        return "daily_convenience"
    if class_no == 41:
        if _contains_any(label_text, ["교육", "훈련"]):
            return "education_kids_pet"
        if _contains_any(label_text, ["출판", "콘텐츠", "미디어"]):
            return "creator_freelance"
        return "culture_travel_sports"
    if class_no == 42:
        if _contains_any(full_text, ["소프트웨어", "saas", "플랫폼", "클라우드"]):
            return "it_platform_app"
        return "creator_freelance"
    if class_no == 43:
        if _contains_any(label_text, ["숙박"]):
            return "daily_convenience"
        return "food_hospitality"
    if class_no == 44:
        if _contains_any(label_text, ["미용", "뷰티", "화장"]):
            return "beauty_care"
        if _contains_any(label_text, ["반려", "수의", "동물"]):
            return "education_kids_pet"
        return "medical_welfare"
    return "misc_services"


def _category_id(kind: str, class_no: int, subgroup_label: str, heading: str) -> str:
    if kind == "goods":
        return _goods_category_id(class_no, subgroup_label, heading)
    return _services_category_id(class_no, subgroup_label, heading)


def _build_class_catalog(rows: list[dict[str, str]]) -> list[dict]:
    class_rows: list[dict] = []
    for row in rows:
        class_no = _parse_class_number(row.get("A", ""))
        heading = _normalize_text(row.get("B", ""))
        if class_no is None or not heading or not 1 <= class_no <= 45:
            continue
        kind = "goods" if class_no <= 34 else "services"
        class_rows.append(
            {
                "kind": kind,
                "nice_class_no": class_no,
                "nice_class_label": format_nice_class(class_no),
                "class_heading": heading,
                "source": "excel",
            }
        )
    return sorted(class_rows, key=lambda item: int(item["nice_class_no"]))


def _empty_group(kind: str, group_id: str) -> dict:
    meta = CATEGORY_META[kind][group_id]
    return {
        "kind": kind,
        "group_id": group_id,
        "group_label": meta["group_label"],
        "group_hint": meta["group_hint"],
        "classes": [],
        "icon": "",
        "source": "excel",
        "subgroups": [],
    }


def _build_group_catalog(rows: list[dict[str, str]]) -> dict[str, list[dict]]:
    groups_by_kind = {
        kind: {meta["group_id"]: _empty_group(kind, meta["group_id"]) for meta in defs}
        for kind, defs in CATEGORY_DEFS.items()
    }

    for row in rows:
        class_no = _parse_class_number(row.get("A", ""))
        heading = _normalize_text(row.get("B", ""))
        if class_no is None or not heading or not 1 <= class_no <= 45:
            continue

        kind = "goods" if class_no <= 34 else "services"
        nice_class_label = format_nice_class(class_no)
        for index, subgroup_label in enumerate(_split_subgroups(heading), start=1):
            group_id = _category_id(kind, class_no, subgroup_label, heading)
            group = groups_by_kind[kind][group_id]
            keywords = _build_keywords(heading, subgroup_label)
            similarity_mapping = _derive_subgroup_similarity_mapping(kind, class_no, subgroup_label, heading)
            group["classes"] = dedupe_ints([*group["classes"], class_no])
            group["subgroups"].append(
                {
                    "kind": kind,
                    "nice_class_no": class_no,
                    "nice_class_label": nice_class_label,
                    "group_id": group_id,
                    "group_label": group["group_label"],
                    "subgroup_id": f"{group_id}_{class_no:02d}_{index:02d}",
                    "subgroup_label": subgroup_label,
                    "keywords": keywords,
                    "similarity_codes": (
                        similarity_mapping["chosen_codes"][:1]
                        if similarity_mapping["chosen_codes"]
                        else similarity_mapping["candidate_codes"][:1]
                    ),
                    "candidate_similarity_codes": similarity_mapping["candidate_codes"],
                    "similarity_match_reason": similarity_mapping["match_reason"],
                    "similarity_match_confidence": similarity_mapping["match_confidence"],
                    "similarity_fallback_used": similarity_mapping["fallback_used"],
                    "nice_classes": [class_no],
                    "class_heading": heading,
                    "source": "excel",
                }
            )

    result: dict[str, list[dict]] = {"goods": [], "services": []}
    for kind, defs in CATEGORY_DEFS.items():
        for meta in defs:
            group = groups_by_kind[kind][meta["group_id"]]
            if not group["subgroups"]:
                continue
            group["classes"] = dedupe_ints(group["classes"])
            group["nice_class_summary"] = format_nice_classes(group["classes"])
            result[kind].append(group)
    return result


def _build_catalog_from_excel() -> tuple[list[dict], dict[str, list[dict]]]:
    rows = _read_excel_rows(EXCEL_SOURCE_PATH)
    return _build_class_catalog(rows), _build_group_catalog(rows)


def export_catalog_cache() -> tuple[list[dict], dict[str, list[dict]]]:
    class_catalog, group_catalog = _build_catalog_from_excel()
    _write_json(CLASS_CATALOG_PATH, class_catalog)
    _write_json(GROUP_CATALOG_PATH, group_catalog)
    load_nice_class_catalog.cache_clear()
    load_nice_group_catalog.cache_clear()
    return class_catalog, group_catalog


def _cache_is_stale() -> bool:
    if not CLASS_CATALOG_PATH.exists() or not GROUP_CATALOG_PATH.exists():
        return True
    dependency_paths = [
        EXCEL_SOURCE_PATH,
        SIMILARITY_CODE_SOURCE_PATH,
        ALIAS_TABLE_PATH,
        LEARNED_MATCHES_PATH,
        Path(__file__),
    ]
    dependency_mtime = max(
        path.stat().st_mtime for path in dependency_paths if path.exists()
    )
    return (
        CLASS_CATALOG_PATH.stat().st_mtime < dependency_mtime
        or GROUP_CATALOG_PATH.stat().st_mtime < dependency_mtime
    )


def ensure_catalog_cache() -> None:
    if _cache_is_stale():
        export_catalog_cache()


@lru_cache(maxsize=1)
def load_nice_class_catalog() -> list[dict]:
    ensure_catalog_cache()
    rows = _load_json(CLASS_CATALOG_PATH)
    return sorted(rows, key=lambda item: int(item["nice_class_no"]))


@lru_cache(maxsize=1)
def load_nice_group_catalog() -> dict[str, list[dict]]:
    ensure_catalog_cache()
    return _load_json(GROUP_CATALOG_PATH)


def get_nice_class_map() -> dict[int, dict]:
    return {int(item["nice_class_no"]): item for item in load_nice_class_catalog()}


def get_groups(kind: str) -> list[dict]:
    return list(load_nice_group_catalog().get(kind, []))


def get_group_cards(kind: str) -> list[dict]:
    cards = []
    for group in get_groups(kind):
        cards.append(
            {
                "group_id": group["group_id"],
                "group_label": group["group_label"],
                "group_hint": group.get("group_hint", ""),
                "nice_class_summary": format_nice_classes(group.get("classes", [])),
            }
        )
    return cards


def find_group(kind: str, group_id: str) -> dict | None:
    for group in get_groups(kind):
        if group.get("group_id") == group_id:
            return group
    return None


def flatten_subgroups(kind: str | None = None) -> list[dict]:
    groups_by_kind = load_nice_group_catalog()
    selected_kinds = [kind] if kind else list(groups_by_kind.keys())
    rows: list[dict] = []
    for current_kind in selected_kinds:
        for group in groups_by_kind.get(current_kind, []):
            for subgroup in group.get("subgroups", []):
                rows.append(
                    {
                        "kind": current_kind,
                        "group_id": group["group_id"],
                        "group_label": group["group_label"],
                        "group_hint": group.get("group_hint", ""),
                        "group_icon": group.get("icon", ""),
                        "group_classes": list(group.get("classes", [])),
                        "nice_class_summary": group.get("nice_class_summary", ""),
                        **subgroup,
                    }
                )
    return rows


def subgroup_to_field(subgroup: dict) -> dict:
    nice_classes = dedupe_ints(subgroup.get("nice_classes", []))
    keywords = dedupe_strings(subgroup.get("keywords", []))
    similarity_codes = dedupe_strings(subgroup.get("similarity_codes", []))
    candidate_similarity_codes = dedupe_strings(
        subgroup.get("candidate_similarity_codes", subgroup.get("similarity_codes", []))
    )
    return {
        "field_id": subgroup["subgroup_id"],
        "kind": subgroup["kind"],
        "group_id": subgroup["group_id"],
        "group_label": subgroup["group_label"],
        "group_hint": subgroup.get("group_hint", ""),
        "description": subgroup["subgroup_label"],
        "example": ", ".join(keywords[:3]),
        "class_no": format_nice_classes(nice_classes),
        "nice_classes": nice_classes,
        "keywords": keywords,
        "similarity_codes": similarity_codes,
        "candidate_similarity_codes": candidate_similarity_codes,
        "similarity_match_reason": subgroup.get("similarity_match_reason", ""),
        "similarity_match_confidence": subgroup.get("similarity_match_confidence", ""),
        "similarity_fallback_used": bool(subgroup.get("similarity_fallback_used", False)),
        "source": subgroup.get("source", "excel"),
    }


def selected_group_labels(selected_fields: Iterable[dict]) -> list[str]:
    return dedupe_strings(field.get("group_label", "") for field in selected_fields)


def selected_group_ids(selected_fields: Iterable[dict]) -> list[str]:
    return dedupe_strings(field.get("group_id", "") for field in selected_fields)


def selected_subgroup_labels(selected_fields: Iterable[dict]) -> list[str]:
    return dedupe_strings(field.get("description", "") for field in selected_fields)


def selected_subgroup_ids(selected_fields: Iterable[dict]) -> list[str]:
    return dedupe_strings(
        field.get("field_id", field.get("subgroup_id", ""))
        for field in selected_fields
    )


def recommended_similarity_codes(selected_fields: Iterable[dict]) -> list[str]:
    return dedupe_strings(
        code
        for field in selected_fields
        for code in field.get("similarity_codes", [])
    )


def can_enter_subgroup_stage(selected_kind: str | None, selected_group: str | None) -> bool:
    return bool(selected_kind and selected_group)


def is_subgroup_selection_complete(selected_fields: Iterable[dict]) -> bool:
    return bool(list(selected_fields))


def can_run_review(selected_subgroup_ids: Iterable[str]) -> bool:
    return bool(dedupe_strings(selected_subgroup_ids))


def can_continue_to_code_selection(selected_fields: Iterable[dict]) -> bool:
    return is_subgroup_selection_complete(selected_fields)


def should_render_subgroup_stage(
    step_scope_sub: str | None,
    selected_kind: str | None,
    selected_group_id: str | None,
) -> bool:
    return (
        step_scope_sub in {SCOPE_SUBSTEP_SUBGROUP, SCOPE_SUBSTEP_REVIEW_READY}
        and can_enter_subgroup_stage(selected_kind, selected_group_id)
    )


def normalize_scope_substep(
    current_substep: str | None,
    selected_group_id: str | None,
    selected_subgroup_ids: Iterable[str],
) -> str:
    if can_run_review(selected_subgroup_ids):
        return SCOPE_SUBSTEP_REVIEW_READY
    if current_substep in {SCOPE_SUBSTEP_SUBGROUP, SCOPE_SUBSTEP_REVIEW_READY} and selected_group_id:
        return SCOPE_SUBSTEP_SUBGROUP
    return SCOPE_SUBSTEP_GROUP


def build_selection_summary(selected_kind: str | None, selected_fields: Iterable[dict]) -> dict:
    fields = list(selected_fields)
    selected_group_id_values = selected_group_ids(fields)
    selected_groups = selected_group_labels(fields)
    selected_subgroup_id_values = selected_subgroup_ids(fields)
    selected_subgroups = selected_subgroup_labels(fields)
    selected_nice_classes = dedupe_ints(
        class_no
        for field in fields
        for class_no in field.get("nice_classes", [])
    )
    return {
        "selected_kind_label": "제품" if selected_kind == "goods" else "서비스" if selected_kind == "services" else "-",
        "selected_group_ids": selected_group_id_values,
        "selected_groups": selected_groups,
        "selected_subgroup_ids": selected_subgroup_id_values,
        "selected_subgroups": selected_subgroups,
        "selected_nice_classes": selected_nice_classes,
        "selected_nice_classes_text": format_nice_classes(selected_nice_classes),
        "recommended_similarity_codes": recommended_similarity_codes(fields),
    }


def derive_selected_scope(
    selected_kind: str | None,
    selected_fields: Iterable[dict],
    specific_products: dict[str, str] | None = None,
    code_lookup: Callable[..., list[dict]] | None = None,
) -> dict:
    fields = list(selected_fields)
    specific_products = specific_products or {}
    summary = build_selection_summary(selected_kind, fields)
    similarity_match_details = [
        {
            "subgroup": field.get("description", ""),
            "candidate_similarity_codes": dedupe_strings(
                field.get("candidate_similarity_codes", field.get("similarity_codes", []))
            ),
            "chosen_similarity_codes": dedupe_strings(field.get("similarity_codes", [])),
            "match_reason": field.get("similarity_match_reason", "catalog_mapping"),
            "match_confidence": field.get("similarity_match_confidence", "high"),
            "fallback_used": bool(field.get("similarity_fallback_used", False)),
        }
        for field in fields
    ]

    subgroup_keywords = dedupe_strings(
        keyword
        for field in fields
        for keyword in field.get("keywords", [])
    )
    derived_similarity_codes = recommended_similarity_codes(fields)
    candidate_similarity_codes = dedupe_strings(
        code
        for field in fields
        for code in field.get("candidate_similarity_codes", field.get("similarity_codes", []))
    )

    if code_lookup is not None:
        for field in fields:
            specific_product = str(specific_products.get(field.get("field_id", ""), "")).strip()
            if not specific_product:
                continue
            subgroup_keywords = dedupe_strings([*subgroup_keywords, specific_product])
            code_rows = code_lookup(
                specific_product,
                limit=10,
                seed_classes=field.get("nice_classes", []),
                seed_keywords=[
                    keyword
                    for keyword in field.get("keywords", [])
                    if ";" not in str(keyword) and len(str(keyword)) <= 40
                ],
                seed_codes=field.get("similarity_codes", []),
            )
            selected_rows = [row for row in code_rows if row.get("selected", True)]
            derived_similarity_codes = dedupe_strings(
                [*derived_similarity_codes, *(row.get("code", "") for row in selected_rows)]
            )
            candidate_similarity_codes = dedupe_strings(
                [*candidate_similarity_codes, *(row.get("code", "") for row in code_rows)]
            )
            if code_rows:
                similarity_match_details.append(
                    {
                        "subgroup": specific_product,
                        "candidate_similarity_codes": dedupe_strings(row.get("code", "") for row in code_rows),
                        "chosen_similarity_codes": dedupe_strings(row.get("code", "") for row in selected_rows),
                        "match_reason": code_rows[0].get("match_reason", "normalized_semantic_match"),
                        "match_confidence": code_rows[0].get("match_confidence", "medium"),
                        "fallback_used": any(bool(row.get("fallback_used", False)) for row in code_rows),
                    }
                )

    search_terms_for_prior_marks = dedupe_strings(
        [
            *summary["selected_subgroups"],
            *subgroup_keywords,
            *(value for value in specific_products.values() if str(value or "").strip()),
        ]
    )

    return {
        "selected_kind": selected_kind,
        "selected_group_ids": summary["selected_group_ids"],
        "selected_groups": summary["selected_groups"],
        "selected_subgroup_ids": summary["selected_subgroup_ids"],
        "selected_subgroups": summary["selected_subgroups"],
        "derived_nice_classes": summary["selected_nice_classes"],
        "derived_similarity_codes": derived_similarity_codes,
        "candidate_similarity_codes": candidate_similarity_codes,
        "similarity_match_details": similarity_match_details,
        "subgroup_keywords": subgroup_keywords,
        "search_terms_for_prior_marks": search_terms_for_prior_marks,
        "selected_scope_summary": {
            **summary,
            "derived_similarity_codes": derived_similarity_codes,
            "candidate_similarity_codes": candidate_similarity_codes,
            "similarity_match_details": similarity_match_details,
            "subgroup_keywords": subgroup_keywords,
            "search_terms_for_prior_marks": search_terms_for_prior_marks,
        },
    }


def build_scope_session_state(
    selected_kind: str | None,
    selected_group_id: str | None = None,
    selected_group_label: str | None = None,
    selected_fields: Iterable[dict] | None = None,
    specific_products: dict[str, str] | None = None,
    code_lookup: Callable[..., list[dict]] | None = None,
    current_substep: str | None = SCOPE_SUBSTEP_GROUP,
) -> dict:
    fields = list(selected_fields or [])
    derived_scope = derive_selected_scope(
        selected_kind=selected_kind,
        selected_fields=fields,
        specific_products=specific_products,
        code_lookup=code_lookup,
    )

    group = find_group(selected_kind, selected_group_id) if selected_kind and selected_group_id else None
    resolved_group_label = selected_group_label or (group.get("group_label") if group else "")
    resolved_group_ids = list(derived_scope.get("selected_group_ids", []))
    resolved_group_labels = list(derived_scope.get("selected_groups", []))

    if selected_group_id and selected_group_id not in resolved_group_ids:
        resolved_group_ids = [selected_group_id, *resolved_group_ids]
    if resolved_group_label and resolved_group_label not in resolved_group_labels:
        resolved_group_labels = [resolved_group_label, *resolved_group_labels]

    resolved_subgroup_ids = list(derived_scope.get("selected_subgroup_ids", []))
    resolved_subgroup_labels = list(derived_scope.get("selected_subgroups", []))

    return {
        "selected_kind": selected_kind,
        "selected_group_id": selected_group_id,
        "selected_group_label": resolved_group_label,
        "selected_group_ids": resolved_group_ids,
        "selected_groups": resolved_group_labels,
        "selected_subgroup_ids": resolved_subgroup_ids,
        "selected_subgroup_labels": resolved_subgroup_labels,
        "selected_subgroups": resolved_subgroup_labels,
        "derived_nice_classes": derived_scope["derived_nice_classes"],
        "derived_similarity_codes": derived_scope["derived_similarity_codes"],
        "candidate_similarity_codes": derived_scope["candidate_similarity_codes"],
        "similarity_match_details": derived_scope["similarity_match_details"],
        "subgroup_keywords": derived_scope["subgroup_keywords"],
        "search_terms_for_prior_marks": derived_scope["search_terms_for_prior_marks"],
        "selected_scope_summary": derived_scope["selected_scope_summary"],
        "step_scope_sub": normalize_scope_substep(
            current_substep=current_substep,
            selected_group_id=selected_group_id,
            selected_subgroup_ids=resolved_subgroup_ids,
        ),
    }


def validate_catalog_coverage() -> dict:
    class_catalog = load_nice_class_catalog()
    class_numbers = {int(item["nice_class_no"]) for item in class_catalog}
    goods_expected = set(range(1, 35))
    services_expected = set(range(35, 46))

    goods_group_classes = set()
    services_group_classes = set()
    empty_subgroup_classes: list[int] = []

    for subgroup in flatten_subgroups("goods"):
        goods_group_classes.update(int(value) for value in subgroup.get("nice_classes", []))
    for subgroup in flatten_subgroups("services"):
        services_group_classes.update(int(value) for value in subgroup.get("nice_classes", []))

    for class_no in goods_expected:
        if not any(int(class_no) in group.get("classes", []) for group in get_groups("goods")):
            empty_subgroup_classes.append(class_no)
    for class_no in services_expected:
        if not any(int(class_no) in group.get("classes", []) for group in get_groups("services")):
            empty_subgroup_classes.append(class_no)

    group_catalog = load_nice_group_catalog()
    return {
        "goods_class_count": len([item for item in class_catalog if item["kind"] == "goods"]),
        "services_class_count": len([item for item in class_catalog if item["kind"] == "services"]),
        "missing_goods_classes": sorted(goods_expected - class_numbers),
        "missing_services_classes": sorted(services_expected - class_numbers),
        "unmapped_goods_classes": sorted(goods_expected - goods_group_classes),
        "unmapped_services_classes": sorted(services_expected - services_group_classes),
        "group_count_goods": len(group_catalog.get("goods", [])),
        "group_count_services": len(group_catalog.get("services", [])),
        "empty_subgroup_classes": sorted(dedupe_ints(empty_subgroup_classes)),
    }
